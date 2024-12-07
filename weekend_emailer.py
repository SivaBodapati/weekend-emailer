# -*- coding: utf-8 -*-
"""Email.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1_swcS7p6uAZnSoGy8cO1q6_icCQGNObe
"""

import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl

def get_weekend_workers(excel_file):
    print("Loading Excel file...")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    weekend_workers = {}
    print("Processing rows in the sheet...")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        date_cell = row[0].value
        day_cell = row[1].value
        emails = row[2].value

        if day_cell in ['Saturday', 'Sunday']:
            if date_cell not in weekend_workers:
                weekend_workers[date_cell] = []
            if emails:
                email_list = emails.split(" - ")
                weekend_workers[date_cell].extend(email_list)

    # Convert the keys to the format '%d-%b-%y' for consistency
    weekend_workers = {
        date.strftime('%d-%b-%y'): emails for date, emails in weekend_workers.items()
    }

    print("Weekend workers extracted:", weekend_workers)
    return weekend_workers

def send_email(recipient, subject, body):
    sender_email = "191fa07001p@gmail.com"
    sender_password = "rdeb bapg abzv jdrb"
    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)
        print(f"Email sent to {recipient}")
    except smtplib.SMTPAuthenticationError:
        print("Authentication failed, check your credentials.")
    except Exception as e:
        print(f"Failed to send email to {recipient}: {e}")

def send_weekend_emails(excel_file):
    today = datetime.date.today()
    print(f"Today is: {today.strftime('%A')} ({today})")

    if today.strftime('%A') == 'Saturday':
        print("It's Saturday, processing emails...")
        weekend_workers = get_weekend_workers(excel_file)

        this_saturday = today
        this_sunday = today + datetime.timedelta(days=1)

        this_saturday_str = this_saturday.strftime('%d-%b-%y')
        this_sunday_str = this_sunday.strftime('%d-%b-%y')

        print(f"This Saturday: {this_saturday_str}, This Sunday: {this_sunday_str}")

        # Combine workers for Saturday and Sunday
        workers_to_email = set(weekend_workers.get(this_saturday_str, []) + weekend_workers.get(this_sunday_str, []))
        print("Workers to email:", workers_to_email)

        if workers_to_email:
            for worker in workers_to_email:
                subject = "Weekend Work Reminder"
                body = f"Dear {worker},\n\nThis is a reminder that you are scheduled to work this weekend. Please be prepared.\n\nBest regards,\nYour Manager"
                print(f"Sending email to {worker}...")
                send_email(worker, subject, body)
        else:
            print("No workers to email today.")
    else:
        print("Not a Saturday")

if __name__ == "__main__":
    excel_file = "Answer.xlsx"  # Replace with your actual file path
    send_weekend_emails(excel_file)

